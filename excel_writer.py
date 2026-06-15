from openpyxl import Workbook
from openpyxl.styles import Font
import os


def write_invoice_to_excel(invoice_data: dict, output_path: str = "output/invoice_output.xlsx"):
    if not invoice_data:
        raise ValueError("Invoice data cannot be empty")

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()
    bold = Font(bold=True)

    # Sheet 1 - Invoice Summary
    ws1 = wb.active
    ws1.title = "Invoice Summary"

    summary_rows = [
        ("Vendor", invoice_data.get("vendor", "")),
        ("Invoice Number", invoice_data.get("invoice_number", "")),
        ("Invoice Date", invoice_data.get("invoice_date", "")),
        ("Due Date", invoice_data.get("due_date", "")),
        ("Subtotal", invoice_data.get("subtotal", 0)),
        ("Tax", invoice_data.get("tax", 0)),
        ("Total", invoice_data.get("total", 0)),
    ]
    for label, value in summary_rows:
        ws1.append([label, value])
        ws1.cell(row=ws1.max_row, column=1).font = bold

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 30

    # Sheet 2 - Line Items
    ws2 = wb.create_sheet("Line Items")
    headers = ["Description", "Quantity", "Price"]
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        ws2.cell(row=1, column=col).font = bold

    for item in invoice_data.get("line_items", []):
        ws2.append([
            item.get("description", ""),
            item.get("quantity", ""),
            item.get("price", ""),
        ])

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    wb.save(output_path)
    return output_path
