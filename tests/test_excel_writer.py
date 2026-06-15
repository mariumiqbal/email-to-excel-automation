import os
import sys

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from excel_writer import write_invoice_to_excel


SAMPLE_DATA = {
    "vendor": "Tech Supplies Inc",
    "invoice_number": "INV-2025-001",
    "invoice_date": "May 1, 2025",
    "due_date": "May 31, 2025",
    "line_items": [
        {"description": "Cloud Storage Service", "quantity": "12 months", "price": 500.00},
        {"description": "Technical Support", "quantity": "10 hours", "price": 750.00},
        {"description": "Software License", "quantity": "1 year", "price": 1200.00},
    ],
    "subtotal": 2450.00,
    "tax": 245.00,
    "total": 2695.00,
}


def test_write_invoice_creates_file(tmp_path):
    output = tmp_path / "out.xlsx"
    result = write_invoice_to_excel(SAMPLE_DATA, str(output))

    assert result == str(output)
    assert output.exists()


def test_write_invoice_sheets_and_content(tmp_path):
    output = tmp_path / "out.xlsx"
    write_invoice_to_excel(SAMPLE_DATA, str(output))

    wb = load_workbook(str(output))
    assert wb.sheetnames == ["Invoice Summary", "Line Items"]

    summary = wb["Invoice Summary"]
    summary_values = {row[0].value: row[1].value for row in summary.iter_rows()}
    assert summary_values["Vendor"] == "Tech Supplies Inc"
    assert summary_values["Invoice Number"] == "INV-2025-001"
    assert summary_values["Total"] == 2695.00

    items = wb["Line Items"]
    assert items.cell(row=1, column=1).value == "Description"
    # Header row + 3 line items
    assert items.max_row == 4
    assert items.cell(row=2, column=1).value == "Cloud Storage Service"


def test_write_invoice_empty_data_raises():
    with pytest.raises(ValueError):
        write_invoice_to_excel({})
